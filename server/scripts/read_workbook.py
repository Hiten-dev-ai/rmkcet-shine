import json
import sys
from datetime import date, datetime, time
from pathlib import Path


def cell_value(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def fill_color(cell):
    fill = getattr(cell, "fill", None)
    if not fill:
        return None
    fg = getattr(fill, "fgColor", None)
    if not fg:
        return None
    rgb = getattr(fg, "rgb", None)
    if isinstance(rgb, str) and rgb and rgb != "00000000":
        return rgb[-6:]
    indexed = getattr(fg, "indexed", None)
    if isinstance(indexed, int):
        return str(indexed)
    return None


def main():
    if len(sys.argv) < 2:
        raise RuntimeError("Usage: read_workbook.py <file_path>")

    file_path = Path(sys.argv[1]).resolve()
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required for workbook parsing.") from exc

    workbook = load_workbook(file_path, data_only=True, read_only=False)
    payload = {"sheetNames": [], "sheets": {}}

    for worksheet in workbook.worksheets:
        sheet_name = worksheet.title
        payload["sheetNames"].append(sheet_name)
        rows = []
        cells = {}
        max_row = worksheet.max_row or 0
        max_column = worksheet.max_column or 0

        for row_index in range(1, max_row + 1):
            row_values = []
            has_value = False
            for col_index in range(1, max_column + 1):
                cell = worksheet.cell(row=row_index, column=col_index)
                value = cell_value(cell.value)
                if value is not None:
                    has_value = True
                row_values.append(value)

                color = fill_color(cell)
                if value is not None or color:
                    key = cell.coordinate
                    item = {"v": value, "w": cell_text(cell.value)}
                    if color:
                        item["s"] = {"fgColor": {"rgb": color}}
                    cells[key] = item

            if has_value:
                while row_values and row_values[-1] is None:
                    row_values.pop()
            rows.append(row_values)

        payload["sheets"][sheet_name] = {
            "rows": rows,
            "cells": cells,
            "maxRow": max_row,
            "maxColumn": max_column,
        }

    sys.stdout.write(json.dumps(payload, ensure_ascii=True))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        sys.stderr.write(str(exc))
        sys.exit(1)
